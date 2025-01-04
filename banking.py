import keyboard as key
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import time
import os
import random
import datetime

# A global cache to store login details during a session
login_cache = {"Account Number":None, "Account Index":None}

class mainMenu:

    def __init__(self):

        # Check if necessary files exist, otherwise create them
        if(os.path.exists("account.xlsx") == False or os.path.exists("transaction.xlsx") == False):
            print("Creating banking enviornment...")
            wb = Workbook()
            wb.save("account.xlsx")
            wb.save("transaction.xlsx")
            wb.close()

            # Set up the account file with header information
            wb = load_workbook("account.xlsx")
            sheet = wb.active
            sheet = wb['Sheet']
            sheet['A1'] = "Account Number"
            sheet['B1'] = "Full Name"
            sheet['C1'] = "Address"
            sheet['D1'] = "Aaddhar Number"
            sheet['E1'] = "Phone Number"
            sheet['F1'] = "Email"
            sheet['G1'] = "Password"
            wb.save("account.xlsx")
            wb.close()

            # Set up the transaction file with header information
            wb = load_workbook("transaction.xlsx")
            sheet = wb.active
            sheet = wb['Sheet']
            sheet['A1'] = "Account Number"
            sheet['B1'] = "Withdrawal/Deposit"
            sheet['C1'] = "Amount"
            sheet['D1'] = "Available"
            sheet['E1'] = "Date/Time"
            wb.save("transaction.xlsx")
            wb.close()
            time.sleep(0.5)
            os.system("cls")

        print("Welcome to Banking System!")
        print("1. Create Account\n2. Login\n3. Exit")
        self.main_menu_options = ["--> Create Account", "--> Login","--> Exit"]
        self.mainmenu()

    # Main menu navigation
    def mainmenu(self):

        self.main_menu_index = -1

        while(True):
            # Navigate the menu using keyboard input
            if key.get_hotkey_name()=='down' and self.main_menu_index!=2:
                os.system("cls")
                print("1. Create Account\n2. Login\n3. Exit\n")
                self.main_menu_index += 1
                print(self.main_menu_options[self.main_menu_index])
                time.sleep(0.2)
            elif key.get_hotkey_name()=='up' and self.main_menu_index!=0:
                os.system("cls")
                self.main_menu_index -= 1
                print("1. Create Account\n2. Login\n3. Exit\n")
                print(self.main_menu_options[self.main_menu_index])
                time.sleep(0.2)
            elif key.get_hotkey_name()=='enter':
                break

        # Call corresponding functions based on user choice
        self.main_menu_functions = [bankAccount().create, authentication().login, exit]
        input()
        self.main_menu_functions[self.main_menu_index]()    
        


# Call corresponding functions based on user choice        
class bankAccount:

    def create(self):
        os.system("cls")
        print("Create Bank Account\n")
        self.user_details = {"Full Name":None, "Address":None, "Aaddhar Number":None, "Phone Number":None, "Email":None, "Transaction Pin (6 digit)":None}  
        self.account_number = ""
        
        # Generate a random 10-digit account number
        for loop_run in range(0, 10):
            self.account_number += str(random.randrange(0, 10))

        # Collect user details
        for detail_input in self.user_details.keys():
            self.user_details[detail_input] = input(f"Enter your {detail_input}: ")

        # Save user details to the account file
        wb = load_workbook("account.xlsx")
        sheet = wb.active
        sheet = wb['Sheet']
        self.empty_cell_index = 1
        while(True):
            if sheet[f"A{self.empty_cell_index}"].value == None:
                break
            self.empty_cell_index += 1
        sheet[f'A{self.empty_cell_index}'] = self.account_number
        sheet[f'B{self.empty_cell_index}'] = self.user_details["Full Name"]
        sheet[f'C{self.empty_cell_index}'] = self.user_details["Address"]
        sheet[f'D{self.empty_cell_index}'] = self.user_details["Aaddhar Number"]
        sheet[f'E{self.empty_cell_index}'] = self.user_details["Phone Number"]
        sheet[f'F{self.empty_cell_index}'] = self.user_details["Email"]
        sheet[f'G{self.empty_cell_index}'] = self.user_details["Transaction Pin (6 digit)"]
        wb.save("account.xlsx")
        wb.close()

        # Log account creation in the transaction file
        date_time = datetime.datetime.now()

        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb['Sheet']
        sheet[f'A{self.empty_cell_index}'] = self.account_number 
        sheet[f'B{self.empty_cell_index}'] = "Account Created"
        sheet[f'C{self.empty_cell_index}'] = "NULL"
        sheet[f'D{self.empty_cell_index}'] = "0"
        sheet[f'E{self.empty_cell_index}'] = f"{date_time.strftime("%x")}-{date_time.strftime("%X")}"
        wb.save("transaction.xlsx")
        wb.close()

        # Create a personalized transaction file for the user
        wb = Workbook()
        wb.save(f"{self.account_number}.xlsx")
        wb.close()

        wb = load_workbook(f"{self.account_number}.xlsx")
        sheet = wb.active
        sheet = wb['Sheet']
        sheet[f'A1'] = "Withdrawal/Deposit"
        sheet[f'B1'] = "Amount"
        sheet[f'C1'] = "Available"
        sheet[f'D1'] = "Date-Time"
        wb.save(f"{self.account_number}.xlsx")
        os.system("cls")
        print(f"\n\nYour account no. is {self.account_number}")
        mainMenu().mainmenu()
        
# Class to handle user authentication

class authentication:


    def login(self):

        os.system("cls")
        print("Bank Login\n")
        self.account_number = input("Enter your Account Number: ")
        self.password = input("Enter your Password: ")
        wb = load_workbook("account.xlsx")
        sheet = wb.active
        sheet = wb['Sheet']
        self.account_column = 1
        self.check_cell = True
        while(self.check_cell):
            if sheet[f"A{self.account_column}"].value == None:
                print("No User with such Account Number")
                time.sleep(1)
                self.login()
                break
            elif sheet[f"A{self.account_column}"].value == self.account_number and sheet[f"G{self.account_column}"].value != self.password:
                print("Password Incorrect")
                time.sleep(1)
                self.login()
                break
            elif sheet[f"A{self.account_column}"].value == self.account_number and sheet[f"G{self.account_column}"].value == self.password:
                login_cache["Account Number"] = self.account_number
                bankService().servicemenu()
                break
            self.account_column += 1
            
# Additional classes and methods for managing deposits, withdrawals, and transfers omitted for brevity

class bankService:
    
    def serviceaccountstatement(self):

        # Clear the console for better readability
        os.system("cls")
    
        # Load the account statement workbook based on account number
        wb = load_workbook(f"{login_cache["Account Number"]}.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        self.account_statement_index = 2
        print("Account Statement:")

        # Iterate through each row until an empty row is found
        while sheet[f"A{self.account_statement_index}"].value != None:

            print(f"Transaction Type : {sheet[f"A{self.account_statement_index}"].value}\nAmount : {sheet[f"B{self.account_statement_index}"].value}\nAvailable : {sheet[f"C{self.account_statement_index}"].value}\nDate-Time : {sheet[f"D{self.account_statement_index}"].value}")
            print("\n")
            self.account_statement_index += 1
        
        wb.close()

        # Return to the service menu
        self.servicemenu()
        


    
    def servicedeposit(self):
        
        # Clear the console for better readability
        os.system("cls")

        # Input deposit amount
        deposit_amount = input("Enter Amount(INR) to deposit: ")
        cell_index = 1
        transact_index = 1
        
        # Open the transaction workbook to update balance
        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        # Find the account in the transaction workbook
        while(True):
            if sheet[f"A{transact_index}"].value == login_cache["Account Number"]:
                break
            transact_index += 1

        available_balance = int(sheet[f"D{transact_index}"].value)
        wb.close()
        
        # Update individual account's statement
        wb = load_workbook(f"{login_cache["Account Number"]}.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        # Find the next empty row
        while(True):
            if sheet[f'A{cell_index}'].value == None:
                break
            cell_index += 1
        
        date_time = datetime.datetime.now()

        # Update transaction details in account statement
        sheet[f"A{cell_index}"] = "Deposited"
        sheet[f"B{cell_index}"] = deposit_amount
        sheet[f"C{cell_index}"] = int(deposit_amount) + available_balance
        sheet[f"D{cell_index}"] = f"{date_time.strftime("%x")}-{date_time.strftime("%x")}"
        wb.save(f"{login_cache["Account Number"]}.xlsx")
        wb.close()

        # Update the transaction workbook with the new balance
        balance_index = 1
        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        while(True):
            if sheet[f'A{balance_index}'].value == login_cache["Account Number"]:
                break
            balance_index += 1
        
        sheet[f"D{balance_index}"] = int(deposit_amount) + available_balance
        wb.save("transaction.xlsx")
        wb.close()

        # Return to the service menu
        self.servicemenu()
        
        
    def servicewithdrawal(self):

        # Clear the console for better readability
        os.system("cls")

        # Input withdrawal amount
        withdrawal_amount = input("Enter Amount(INR) to Withdraw: ")
        cell_index = 1
        transact_index = 1
        
        # Open the transaction workbook to fetch available balance
        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        # Locate the account in the transaction workbook
        while(True):
            if sheet[f"A{transact_index}"].value == login_cache["Account Number"]:
                break
            transact_index += 1

        available_balance = int(sheet[f"D{transact_index}"].value)
        wb.close()
        
        # Check if sufficient balance is available
        if available_balance >= int(withdrawal_amount):
            wb = load_workbook(f"{login_cache["Account Number"]}.xlsx")
            sheet = wb.active
            sheet = wb["Sheet"]

            # Locate the next empty row in the account statement
            while(True):
                if sheet[f'A{cell_index}'].value == None:
                    break
                cell_index += 1

            date_time = datetime.datetime.now()

            # Update the withdrawal details in the account statement
            sheet[f"A{cell_index}"] = "Withdrawed"
            sheet[f"B{cell_index}"] = withdrawal_amount
            sheet[f"C{cell_index}"] = available_balance - int(withdrawal_amount)
            sheet[f"D{cell_index}"] = f"{date_time.strftime("%x")}-{date_time.strftime("%x")}"
            wb.save(f"{login_cache["Account Number"]}.xlsx")
            wb.close()

            # Update the transaction workbook with the new balance
            balance_index = 1
            wb = load_workbook("transaction.xlsx")
            sheet = wb.active
            sheet = wb["Sheet"]

            while(True):
                if sheet[f'A{balance_index}'].value == login_cache["Account Number"]:
                    break
                balance_index += 1

            sheet[f"D{balance_index}"] = available_balance - int(withdrawal_amount) 
            wb.save("transaction.xlsx")
            wb.close()
        
        else:
            print("Not Enough Available Blance")
            self.servicemenu()
        
        # Return to the service menu
        self.servicemenu()

    def transfer(self):

        # Set the account index for login cache
        login_cache["Account Index"] = self.transact_index

        account_check_index = 1
        
        # Open the transaction workbook
        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        # Verify the account to transfer to exists
        while(True):
            if sheet[f"A{account_check_index}"].value == self.account_transfer_number:
                break
            elif sheet[f"A{account_check_index}"].value == None:
                print("No Account found")
                self.servicemenu()
            account_check_index += 1
        
        # Ensure sufficient balance for the transfer
        if int(sheet[f"D{login_cache["Account Index"]}"].value) >= int(sheet[f"D{account_check_index}"].value):
            
            # Deduct amount from sender's account
            sheet[f"D{login_cache["Account Index"]}"] = int(sheet[f"D{login_cache["Account Index"]}"].value) - int(self.transfer_amount)

            # Add amount to receiver's account
            sheet[f"D{account_check_index}"] = int(sheet[f"D{account_check_index}"].value) + int(self.transfer_amount)
        wb.save("transaction.xlsx")
        wb.close()

        # Update sender's transaction statement
        wb = load_workbook(f"{login_cache["Account Number"]}.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        cell_index = 1

        # Locate next empty row in sender's account statement
        while(True):
            if sheet[f'A{cell_index}'].value == None:
                break
            cell_index += 1

        date_time = datetime.datetime.now()
        sheet[f"A{cell_index}"] = "Credited"
        sheet[f"B{cell_index}"] = self.transfer_amount
        sheet[f"C{cell_index}"] = int(sheet[f"C{(login_cache["Account Index"])}"].value) - int(self.transfer_amount)
        sheet[f"D{cell_index}"] = f"{date_time.strftime("%x")}-{date_time.strftime("%x")}"
        
        wb.save(f"{login_cache["Account Number"]}.xlsx")
        wb.close()

        # Update receiver's transaction statement
        wb = load_workbook(f"{self.account_transfer_number}.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        cell_index_2 = 1

        # Locate next empty row in receiver's account statement
        while(True):
            if sheet[f'A{cell_index_2}'].value == None:
                break
            cell_index_2 += 1

        date_time = datetime.datetime.now()
        sheet[f"A{cell_index_2}"] = "Deposited"
        sheet[f"B{cell_index_2}"] = self.transfer_amount
        try:
            sheet[f"C{cell_index_2}"] = int(sheet[f"C{login_cache["Account Index"]}"].value) + self.transfer_amount
        except:
            sheet[f"C{cell_index_2}"] = 0 + int(self.transfer_amount)         
        sheet[f"D{cell_index_2}"] = f"{date_time.strftime("%x")}-{date_time.strftime("%x")}"
        
        wb.save(f"{self.account_transfer_number}.xlsx")
        wb.close()

        # Return to the service menu
        self.servicemenu()


    def serviceaccounttransfer(self):
        
        # Clear the console for better readability
        os.system("cls")

        # Input receiver's account number and transfer amount
        self.account_transfer_number = input("Enter person's Account Number: ")
        self.transfer_amount = input("Enter amount to tranfer(INR): ")

        self.transact_index = 1

        # Open the transaction workbook to locate sender's account
        wb = load_workbook("transaction.xlsx")
        sheet = wb.active
        sheet = wb["Sheet"]

        while(True):
            if sheet[f"A{self.transact_index}"].value == login_cache["Account Number"]:
                break
            self.transact_index += 1
        wb.close()
        
        # Proceed with transfer
        self.transfer()

        
    def service_2(self):

        key = None
        
        # Handle menu navigation and selection
        while(key!="enter"):
            key = input("Enter a Key: up, down, enter: ")
            if key.lower() == 'down' and self.service_menu_index!=3:
                os.system("cls")
                print("1. Deposit\n2. Withdrawal\n3. Account transfer\n4. Account Statement\n5. Back")
                self.service_menu_index += 1
                print(self.service_menu_options[self.service_menu_index])
                time.sleep(0.2)

            elif key.lower() == 'up' and self.service_menu_index!=0:
                os.system("cls")
                self.service_menu_index -= 1
                print("1. Deposit\n2. Withdrawal\n3. Account transfer\n4. Account Statement\n5. Back")
                print(self.service_menu_options[self.service_menu_index])
                time.sleep(0.2)

            elif key.lower() == 'enter':
                break


        # Call the selected service menu function
        self.service_menu_functions = [self.servicedeposit, self.servicewithdrawal, self.serviceaccounttransfer, self.serviceaccountstatement, exit]
        input()
        self.service_menu_functions[self.service_menu_index]()



    def servicemenu(self):

        # Display the main service menu
        print("1. Deposit\n2. Withdrawal\n3. Account transfer\n4. Account Statement\n5. Back")
        self.service_menu_index = -1
        self.service_menu_options = ["--> Deposit", "--> Withdrawal", "--> Account transfer", "--> Account Statement", "--> Back"]
        self.service_2()
        
        
        
    

# Initialize the banking system

bank_exe = mainMenu()



